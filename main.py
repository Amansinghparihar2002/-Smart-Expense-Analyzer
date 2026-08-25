# app.py
# Smart Expense Analyzer — with Gemini AI Financial Advisor Chatbot
# Run with: streamlit run app.py

import streamlit as st
import pandas as pd
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from data_processing import clean_data, add_time_features, apply_filters
from charts import (
    plot_category_bar,
    plot_category_pie,
    plot_monthly_trend,
    plot_budget_comparison,
    plot_top_transactions,
)
from insights import generate_insights, compute_health_score
from chatbot import get_chatbot_response, get_suggested_questions

# ─────────────────────────────────────────────────────────────────────────────
# YOUR LOCAL CSV PATH
# ─────────────────────────────────────────────────────────────────────────────
DEFAULT_CSV_PATH = r"C:\Datascience study material\ML\Expense analyzer\Credit card transactions.csv"

# REPORT GENERATION FUNCTION

def create_expense_report(
    total_spend,
    total_txns,
    avg_txn,
    top_cat,
    top_cat_amt,
    score,
    label,
    category_summary,
    monthly_summary,
    insights
):
    wb = Workbook()

    # Remove default sheet
    ws = wb.active
    ws.title = "Expense Report"

    # Title
    ws["A1"] = "Smart Expense Analyzer - Expense Report"
    ws["A1"].font = Font(size=18, bold=True)
    ws.merge_cells("A1:B1")

    # Overall Summary
    ws["A3"] = "Overall Summary"
    ws["A3"].font = Font(size=14, bold=True)

    summary_data = [
        ["Total Spending", total_spend],
        ["Number of Transactions", total_txns],
        ["Average Transaction", avg_txn],
        ["Top Spending Category", top_cat],
        ["Top Category Spending", top_cat_amt],
        ["Financial Health Score", f"{score}/100"],
        ["Financial Health Label", label],
    ]

    row = 4
    for key, value in summary_data:
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Category-wise spending
    ws = wb.create_sheet("Category Spending")

    ws["A1"] = "Category-wise Spending"
    ws["A1"].font = Font(size=14, bold=True)

    headers = ["Category", "Total Spend", "Transactions", "Average Amount", "% of Total"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)

    row = 4

    for category, data in category_summary.items():
        ws.cell(row=row, column=1, value=category)
        ws.cell(row=row, column=2, value=data["total_spend"])
        ws.cell(row=row, column=3, value=data["transactions"])
        ws.cell(row=row, column=4, value=data["avg_amount"])
        ws.cell(row=row, column=5, value=data["percentage"])
        row += 1

    # Monthly spending
    ws = wb.create_sheet("Monthly Spending")

    ws["A1"] = "Monthly Spending"
    ws["A1"].font = Font(size=14, bold=True)

    headers = ["Month", "Total Spend", "Transactions"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)

    row = 4

    for month, data in monthly_summary.items():
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=data["total_spend"])
        ws.cell(row=row, column=3, value=data["transactions"])
        row += 1

    # Smart Insights
    ws = wb.create_sheet("Smart Insights")

    ws["A1"] = "Smart Insights"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = "Type"
    ws["B3"] = "Insight"

    ws["A3"].font = Font(bold=True)
    ws["B3"].font = Font(bold=True)

    row = 4

    for insight in insights:
        ws.cell(row=row, column=1, value=insight["type"])
        ws.cell(row=row, column=2, value=insight["message"])
        row += 1

    # Formatting
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            sheet.column_dimensions[column_letter].width = min(max_length + 3, 50)

        for cell in sheet[1]:
            cell.alignment = Alignment(vertical="center")

    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

# ─────────────────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Smart Expense Analyzer",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Compute this early — used in the header status pill and again in the AI tab.
api_key_set = bool(os.getenv("GEMINI_API_KEY"))

# ─────────────────────────────────────────────────────────────────────────────
# Custom CSS — "Premium FinTech Analytics Dashboard" theme
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    :root {
        --navy-950: #0a0e1f;
        --navy-900: #0f1530;
        --navy-800: #16204a;
        --navy-700: #1c2b5e;
        --accent: #5b7fff;
        --accent-soft: #eef1ff;
        --text-dark: #101323;
        --text-muted: #6b7280;
        --border-soft: #eaecf3;
        --radius-lg: 18px;
        --radius-md: 14px;
        --shadow-card: 0 4px 18px rgba(16, 19, 35, 0.06);
        --shadow-card-hover: 0 10px 28px rgba(16, 19, 35, 0.10);
    }

    html, body, [class*="css"] {
        font-family: "Inter", "Segoe UI", -apple-system, BlinkMacSystemFont, sans-serif;
    }

    .block-container { padding-top: 1.4rem; }

    /* ── Header ─────────────────────────────────────────────────────────── */
    .app-header {
        background: linear-gradient(120deg, var(--navy-950) 0%, var(--navy-900) 55%, var(--navy-700) 100%);
        border-radius: var(--radius-lg);
        padding: 1.5rem 2rem;
        margin-bottom: 1.6rem;
        display: flex;
        align-items: center;
        justify-content: space-between;
        box-shadow: var(--shadow-card);
    }
    .app-header h1 {
        color: #ffffff;
        margin: 0;
        font-size: 1.9rem;
        font-weight: 800;
        letter-spacing: -0.02em;
    }
    .app-header p {
        color: #aab4d6;
        margin: 0.25rem 0 0;
        font-size: 0.95rem;
    }
    .status-pill {
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: rgba(255,255,255,0.06);
        border: 1px solid rgba(255,255,255,0.12);
        border-radius: 999px;
        padding: 6px 14px;
        font-size: 0.8rem;
        color: #e4e8fb;
        white-space: nowrap;
    }
    .status-dot { width: 8px; height: 8px; border-radius: 50%; display: inline-block; }
    .status-on  { background: #2ecc71; box-shadow: 0 0 8px #2ecc71; }
    .status-off { background: #e74c3c; box-shadow: 0 0 8px #e74c3c; }

    /* ── Section headers ────────────────────────────────────────────────── */
    .section-title {
        font-size: 1.05rem;
        font-weight: 700;
        color: var(--text-dark);
        display: flex;
        align-items: center;
        gap: 8px;
        margin: 1.8rem 0 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 1px solid var(--border-soft);
    }

    /* ── KPI cards ──────────────────────────────────────────────────────── */
    .kpi-grid {
        display: grid;
        grid-template-columns: repeat(5, 1fr);
        gap: 14px;
        margin-bottom: 0.5rem;
    }
    @media (max-width: 1100px) { .kpi-grid { grid-template-columns: repeat(2, 1fr); } }
    .kpi-card {
        background: #ffffff;
        border: 1px solid var(--border-soft);
        border-radius: var(--radius-md);
        padding: 1rem 1.1rem;
        box-shadow: var(--shadow-card);
        transition: box-shadow 0.2s ease, transform 0.2s ease;
    }
    .kpi-card:hover { box-shadow: var(--shadow-card-hover); transform: translateY(-2px); }
    .kpi-icon { font-size: 1.3rem; margin-bottom: 6px; }
    .kpi-label { font-size: 0.78rem; color: var(--text-muted); font-weight: 600; text-transform: uppercase; letter-spacing: 0.02em; }
    .kpi-value { font-size: 1.5rem; font-weight: 800; color: var(--text-dark); margin-top: 2px; line-height: 1.2; }
    .kpi-sub { font-size: 0.78rem; color: var(--text-muted); margin-top: 4px; }

    /* ── Health score card ─────────────────────────────────────────────── */
    .health-wrap {
        background: #ffffff;
        border: 1px solid var(--border-soft);
        border-radius: var(--radius-lg);
        padding: 1.4rem 1.6rem;
        box-shadow: var(--shadow-card);
        display: flex;
        align-items: center;
        gap: 1.6rem;
    }
    .health-ring {
        width: 130px; height: 130px; border-radius: 50%;
        display: flex; align-items: center; justify-content: center;
        flex-shrink: 0;
    }
    .health-ring-inner {
        width: 100px; height: 100px; border-radius: 50%;
        background: #ffffff;
        display: flex; flex-direction: column; align-items: center; justify-content: center;
    }
    .health-ring-score { font-size: 1.9rem; font-weight: 800; color: var(--text-dark); line-height: 1; }
    .health-ring-max { font-size: 0.7rem; color: var(--text-muted); }
    .health-label-badge {
        display: inline-block; padding: 3px 12px; border-radius: 999px;
        font-size: 0.8rem; font-weight: 700; color: #ffffff; margin-bottom: 6px;
    }
    .health-reason { font-size: 0.88rem; margin: 4px 0; color: var(--text-dark); }

    /* ── Insight cards ──────────────────────────────────────────────────── */
    .insight-card {
        display: flex; gap: 10px; align-items: flex-start;
        border-radius: 12px; padding: 0.7rem 1rem; margin: 0.4rem 0;
        font-size: 0.9rem; line-height: 1.5; border: 1px solid transparent;
    }
    .insight-warning { background: #fff7ed; border-color: #fde3c2; }
    .insight-tip     { background: #f5f0ff; border-color: #e3d6ff; }
    .insight-success { background: #edfcf3; border-color: #c8f2d9; }
    .insight-info    { background: var(--accent-soft); border-color: #dbe2ff; }
    .insight-badge   { font-size: 1rem; flex-shrink: 0; }

    /* ── Suggestion pill buttons ───────────────────────────────────────── */
    .stButton > button {
        border-radius: 999px !important;
        border: 1px solid var(--border-soft) !important;
        font-weight: 600 !important;
        transition: all 0.15s ease !important;
    }
    .stButton > button:hover {
        border-color: var(--accent) !important;
        color: var(--accent) !important;
        box-shadow: 0 4px 12px rgba(91,127,255,0.18) !important;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(120deg, var(--navy-900), var(--accent)) !important;
        border: none !important;
    }

    /* ── Chat ───────────────────────────────────────────────────────────── */
    .chat-container {
        background: #f9fafc;
        border-radius: var(--radius-lg);
        border: 1px solid var(--border-soft);
        padding: 1.1rem;
        max-height: 480px;
        overflow-y: auto;
        margin-bottom: 0.9rem;
    }
    .chat-row { display: flex; flex-direction: column; margin: 0.5rem 0; }
    .chat-row.user { align-items: flex-end; }
    .chat-row.bot  { align-items: flex-start; }
    .chat-name {
        font-size: 0.72rem; font-weight: 700; color: var(--text-muted);
        margin: 0 4px 3px;
    }
    .chat-user {
        background: linear-gradient(120deg, var(--navy-900), var(--accent));
        color: white;
        padding: 0.65rem 1rem;
        border-radius: 16px 16px 4px 16px;
        max-width: 78%;
        font-size: 0.92rem;
        line-height: 1.5;
    }
    .chat-bot {
        background: #ffffff;
        color: var(--text-dark);
        padding: 0.65rem 1rem;
        border-radius: 16px 16px 16px 4px;
        border: 1px solid var(--border-soft);
        max-width: 78%;
        font-size: 0.92rem;
        line-height: 1.5;
        box-shadow: var(--shadow-card);
    }
    .chat-empty { text-align: center; color: var(--text-muted); padding: 2.2rem; font-size: 0.9rem; }

    /* ── Generic card wrapper used around charts/tables ───────────────── */
    .panel-caption { color: var(--text-muted); font-size: 0.85rem; margin-top: -0.3rem; margin-bottom: 0.6rem; }

    /* ── Sidebar ────────────────────────────────────────────────────────── */
    section[data-testid="stSidebar"] {
        background: var(--navy-950);
    }
    section[data-testid="stSidebar"] * { color: #dfe3f5 !important; }
    section[data-testid="stSidebar"] .stTextInput input,
    section[data-testid="stSidebar"] .stDateInput input {
        background: rgba(255,255,255,0.06) !important;
        color: #ffffff !important;
        border: 1px solid rgba(255,255,255,0.12) !important;
    }
    .sidebar-section-label {
        font-size: 0.72rem; letter-spacing: 0.08em; font-weight: 700;
        color: #7c88b8 !important; text-transform: uppercase;
        margin: 1.1rem 0 0.4rem;
    }

    /* ── Tabs ───────────────────────────────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] { gap: 6px; border-bottom: 1px solid var(--border-soft); }
    .stTabs [data-baseweb="tab"] {
        border-radius: 10px 10px 0 0 !important;
        padding: 8px 18px !important;
        font-weight: 600 !important;
    }
    .stTabs [aria-selected="true"] {
        background: var(--accent-soft) !important;
        color: var(--accent) !important;
    }

    /* ── Landing page hero ─────────────────────────────────────────────── */
    .hero-card {
        background: linear-gradient(120deg, var(--navy-950) 0%, var(--navy-900) 55%, var(--navy-700) 100%);
        border-radius: var(--radius-lg);
        padding: 2.4rem 2.2rem;
        color: #ffffff;
        margin-bottom: 1.6rem;
    }
    .hero-card h1 { font-size: 2.1rem; margin: 0 0 0.6rem; font-weight: 800; }
    .hero-card p  { color: #aab4d6; font-size: 1.02rem; line-height: 1.6; margin: 0; }
    .feature-card {
        background: #ffffff; border: 1px solid var(--border-soft);
        border-radius: var(--radius-md); padding: 1rem 1.1rem;
        box-shadow: var(--shadow-card); height: 100%;
    }
    .feature-card h4 { margin: 0 0 4px; font-size: 0.98rem; color: var(--text-dark); }
    .feature-card p  { margin: 0; font-size: 0.84rem; color: var(--text-muted); line-height: 1.45; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Header
# ─────────────────────────────────────────────────────────────────────────────
status_class = "status-on" if api_key_set else "status-off"
status_text  = "AI Advisor Online" if api_key_set else "AI Advisor Offline"

st.markdown(f"""
<div class="app-header">
    <div>
        <h1>💰 Smart Expense Analyzer</h1>
        <p>AI-powered personal finance intelligence</p>
    </div>
    <div class="status-pill"><span class="status-dot {status_class}"></span>{status_text}</div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Session state — chat history persists across Streamlit reruns
# ─────────────────────────────────────────────────────────────────────────────
if "chat_messages" not in st.session_state:
    st.session_state.chat_messages = []   # list of {"role": "user"/"bot", "text": "..."}

if "gemini_history" not in st.session_state:
    st.session_state.gemini_history = []  # multi-turn history for Gemini API

# ─────────────────────────────────────────────────────────────────────────────
# Sidebar — Load data
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 💰 Control Panel")

    st.markdown('<div class="sidebar-section-label">Data Source</div>', unsafe_allow_html=True)

    st.markdown("**Local CSV Path**")
    local_path = st.text_input("CSV Path", value=DEFAULT_CSV_PATH, label_visibility="collapsed")
    load_btn   = st.button("📂 Load from Path", use_container_width=True, type="primary")

    st.markdown("**Upload CSV**")
    uploaded = st.file_uploader("Choose CSV", type=["csv"], label_visibility="collapsed")

    st.caption("Expected columns: Date · Exp Type · Amount")

# ─────────────────────────────────────────────────────────────────────────────
# Load data — persist across Streamlit reruns
# ─────────────────────────────────────────────────────────────────────────────

if "raw_df" not in st.session_state:
    st.session_state.raw_df = None

if load_btn:
    path = local_path.strip()

    if not path:
        st.error("Please enter a file path.")
        st.stop()

    if not os.path.exists(path):
        st.error(f"File not found:\n`{path}`\nPlease check the path.")
        st.stop()

    try:
        st.session_state.raw_df = pd.read_csv(path)
        st.sidebar.success("Loaded from local path!")
    except Exception as e:
        st.error(f"Could not read file: {e}")
        st.stop()

elif uploaded is not None:
    try:
        st.session_state.raw_df = pd.read_csv(uploaded)
        st.sidebar.success("File uploaded!")
    except Exception as e:
        st.error(f"Could not read uploaded file: {e}")
        st.stop()

# Use the persisted dataframe
raw_df = st.session_state.raw_df
# ─────────────────────────────────────────────────────────────────────────────
# Process + display dashboard
# ─────────────────────────────────────────────────────────────────────────────
if raw_df is not None:
    try:
        df_clean = clean_data(raw_df)
        df_full  = add_time_features(df_clean)
    except ValueError as ve:
        st.error(f"Data Error: {ve}")
        st.info("Your CSV must have columns: **Date**, **Exp Type**, **Amount**.")
        st.stop()
    except Exception as e:
        st.error(f"Unexpected processing error: {e}")
        st.stop()

    if df_full.empty:
        st.warning("No valid rows found after cleaning. Please check your CSV.")
        st.stop()

    # ── Sidebar filters ───────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown('<div class="sidebar-section-label">Filters</div>', unsafe_allow_html=True)

        all_cats = sorted(df_full["category"].unique().tolist())
        sel_cats = st.multiselect(
            "Categories", options=all_cats, default=all_cats,
            help="Deselect a category to hide it.",
        )

        min_date = df_full["date"].min().date()
        max_date = df_full["date"].max().date()
        date_rng = st.date_input(
            "Date Range", value=(min_date, max_date),
            min_value=min_date, max_value=max_date,
        )
        st.caption("Charts and insights update instantly.")

    # Handle single-date edge case
    if isinstance(date_rng, (list, tuple)) and len(date_rng) == 2:
        start_d, end_d = date_rng
    else:
        start_d = end_d = date_rng if not isinstance(date_rng, (list, tuple)) else date_rng[0]

    df = apply_filters(df_full, sel_cats, (start_d, end_d))

    if df.empty:
        st.warning("No data matches the selected filters.")
        st.stop()

    # Active filter note
    filtered_note = []
    if set(sel_cats) != set(all_cats):
        filtered_note.append(f"{len(sel_cats)} of {len(all_cats)} categories")
    if start_d != min_date or end_d != max_date:
        filtered_note.append(f"{start_d} → {end_d}")
    if filtered_note:
        st.info(f"🔍 Active filters: {' | '.join(filtered_note)}  —  {len(df):,} transactions")

    # ── Pre-compute values used across sections ───────────────────────────────
    total_spend = df["amount"].sum()
    total_txns  = len(df)
    top_cat     = df.groupby("category")["amount"].sum().idxmax()
    top_cat_amt = df.groupby("category")["amount"].sum().max()
    avg_txn     = df["amount"].mean()
    n_months    = df["month_year"].nunique()

    insights        = generate_insights(df)
    score, label, color, reasons = compute_health_score(df)

    # ── Build the summary_stats dict used by chatbot ──────────────────────────
    # ── Build detailed data context used by chatbot ─────────────────────────────

    # Category-level statistics
    category_summary = (
        df.groupby("category")["amount"]
        .agg(
            total_spend="sum",
            transactions="count",
            avg_amount="mean"
        )
        .sort_values("total_spend", ascending=False)
    )

    category_data = {}

    for category, row in category_summary.iterrows():
        category_data[category] = {
            "total_spend": float(row["total_spend"]),
            "transactions": int(row["transactions"]),
            "avg_amount": float(row["avg_amount"]),
            "percentage": float(
                (row["total_spend"] / total_spend) * 100
            ),
        }

    # Monthly-level statistics
    monthly_summary = (
        df.groupby("month_year")["amount"]
        .agg(
            total_spend="sum",
            transactions="count"
        )
        .sort_index()
    )

    monthly_data = {}

    for month, row in monthly_summary.iterrows():
        monthly_data[month] = {
            "total_spend": float(row["total_spend"]),
            "transactions": int(row["transactions"]),
        }

    # Overall summary
    summary_stats = {
        "total_spend": float(total_spend),
        "n_txns": int(total_txns),
        "avg_txn": float(avg_txn),
        "top_cat": top_cat,
        "n_months": int(n_months),
        "health_score": int(score),
        "health_label": label,

        # NEW detailed information
        "category_data": category_data,
        "monthly_data": monthly_data,
    }
    # =========================================================================
    # TAB LAYOUT — Dashboard | AI Advisor
    # =========================================================================
    tab_dash, tab_chat = st.tabs(["📊 Dashboard", "🤖 AI Financial Advisor"])

    # =========================================================================
    # TAB 1 — DASHBOARD
    # =========================================================================
    with tab_dash:

        # ── Key Metrics (premium KPI cards) ─────────────────────────────────
        st.markdown('<div class="section-title">📊 Key Metrics</div>', unsafe_allow_html=True)

        st.markdown(f"""
        <div class="kpi-grid">
            <div class="kpi-card">
                <div class="kpi-icon">💸</div>
                <div class="kpi-label">Total Spend</div>
                <div class="kpi-value">₹{total_spend:,.0f}</div>
                <div class="kpi-sub">across all filtered transactions</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-icon">🔢</div>
                <div class="kpi-label">Transactions</div>
                <div class="kpi-value">{total_txns:,}</div>
                <div class="kpi-sub">total records</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-icon">🏆</div>
                <div class="kpi-label">Top Category</div>
                <div class="kpi-value" style="font-size:1.15rem;">{top_cat}</div>
                <div class="kpi-sub">₹{top_cat_amt:,.0f} spent</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-icon">📈</div>
                <div class="kpi-label">Avg Transaction</div>
                <div class="kpi-value">₹{avg_txn:,.0f}</div>
                <div class="kpi-sub">per transaction</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-icon">📅</div>
                <div class="kpi-label">Months Covered</div>
                <div class="kpi-value">{n_months}</div>
                <div class="kpi-sub">unique months in view</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Financial Health Score ────────────────────────────────────────────
        st.markdown('<div class="section-title">❤️ Financial Health Score</div>', unsafe_allow_html=True)

        with st.container(border=True):
            col_score, col_reasons = st.columns([1, 2])

            with col_score:
                ring_pct = max(0, min(100, score))
                st.markdown(f"""
                <div class="health-wrap" style="border:none; box-shadow:none; padding:0;">
                    <div class="health-ring" style="background: conic-gradient({color} {ring_pct}%, #eef0f5 {ring_pct}% 100%);">
                        <div class="health-ring-inner">
                            <div class="health-ring-score">{score}</div>
                            <div class="health-ring-max">/ 100</div>
                        </div>
                    </div>
                    <div>
                        <span class="health-label-badge" style="background:{color};">{label}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            with col_reasons:
                st.markdown("**Score Breakdown**")
                for r in reasons:
                    icon = "🟢" if r.startswith("+") else "🔴"
                    st.markdown(f'<div class="health-reason">{icon} {r}</div>', unsafe_allow_html=True)
                if not reasons:
                    st.success("No penalties — well balanced!")
                if label == "Poor":
                    st.error("Spending habits need immediate attention.")
                elif label == "Average":
                    st.warning("Room for improvement.")
                elif label == "Good":
                    st.info("Doing well — a few tweaks could push it higher.")
                else:
                    st.success("Excellent financial habits! Keep it up.")

        # ── Charts ────────────────────────────────────────────────────────────
        st.markdown('<div class="section-title">📊 Spending Overview</div>', unsafe_allow_html=True)
        with st.container(border=True):
            col_bar, col_pie = st.columns(2)
            with col_bar:
                st.pyplot(plot_category_bar(df))
            with col_pie:
                st.pyplot(plot_category_pie(df))

        st.markdown('<div class="section-title">📈 Spending Trends</div>', unsafe_allow_html=True)
        with st.container(border=True):
            st.pyplot(plot_monthly_trend(df))

        st.markdown('<div class="section-title">🎯 Budget & Transactions</div>', unsafe_allow_html=True)
        with st.container(border=True):
            col_bud, col_top = st.columns(2)
            with col_bud:
                st.pyplot(plot_budget_comparison(df))
            with col_top:
                st.pyplot(plot_top_transactions(df, n=min(10, len(df))))

        # ── Smart Insights ────────────────────────────────────────────────────
        st.markdown('<div class="section-title">💡 Smart Insights</div>', unsafe_allow_html=True)

        TYPE_ORDER  = ["warning", "tip", "success", "info"]
        TYPE_LABELS = {
            "warning": "⚠️ Warnings & Alerts",
            "tip":     "💡 Savings Tips",
            "success": "✅ Positive Signs",
            "info":    "ℹ️ Summary",
        }
        CSS_CLASS = {
            "warning": "insight-warning",
            "tip":     "insight-tip",
            "success": "insight-success",
            "info":    "insight-info",
        }
        BADGE_ICON = {
            "warning": "⚠️",
            "tip":     "💡",
            "success": "✅",
            "info":    "ℹ️",
        }

        with st.container(border=True):
            for t in TYPE_ORDER:
                bucket = [i for i in insights if i["type"] == t]
                if not bucket:
                    continue
                st.markdown(f"**{TYPE_LABELS[t]}**")
                for item in bucket:
                    st.markdown(
                        f'<div class="insight-card {CSS_CLASS[t]}">'
                        f'<span class="insight-badge">{BADGE_ICON[t]}</span>'
                        f'<span>{item["message"]}</span></div>',
                        unsafe_allow_html=True,
                    )
                st.markdown("")

        # ── Category Summary Table ────────────────────────────────────────────
        st.markdown('<div class="section-title">📋 Category Summary</div>', unsafe_allow_html=True)

        summary = (
            df.groupby("category")["amount"]
            .agg(Total_Spend="sum", Transactions="count", Avg_Amount="mean")
            .sort_values("Total_Spend", ascending=False)
            .reset_index()
        )
        summary["% of Total"] = (summary["Total_Spend"] / total_spend * 100).round(1).astype(str) + "%"
        summary["Total_Spend"] = summary["Total_Spend"].apply(lambda x: f"₹{x:,.0f}")
        summary["Avg_Amount"]  = summary["Avg_Amount"].apply(lambda x: f"₹{x:,.0f}")
        summary.columns        = ["Category", "Total Spend", "Transactions", "Avg Amount", "% of Total"]

        with st.container(border=True):
            st.dataframe(summary, use_container_width=True, hide_index=True)

        # ── Download Report ──────────────────────────────────────────────────
        st.markdown(
            '<div class="section-title">📥 Download Report</div>',
            unsafe_allow_html=True
        )

        report_file = create_expense_report(
            total_spend=total_spend,
            total_txns=total_txns,
            avg_txn=avg_txn,
            top_cat=top_cat,
            top_cat_amt=top_cat_amt,
            score=score,
            label=label,
            category_summary=category_data,
            monthly_summary=monthly_data,
            insights=insights
        )

        st.download_button(
            label="📥 Download Expense Report",
            data=report_file,
            file_name="Smart_Expense_Analyzer_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

        # ── Raw Data Explorer ─────────────────────────────────────────────────
        with st.expander("🗂️ Explore Raw Transactions"):
            sort_col = st.selectbox("Sort by", ["date", "amount", "category"], index=1)
            sort_asc = st.checkbox("Ascending", value=False)
            view_df  = (
                df[["date", "category", "amount"]]
                .sort_values(sort_col, ascending=sort_asc)
                .reset_index(drop=True)
            )
            st.dataframe(view_df, use_container_width=True)
            st.caption(f"{len(view_df):,} rows shown.")

    # =========================================================================
    # TAB 2 — AI FINANCIAL ADVISOR CHATBOT
    # =========================================================================
    with tab_chat:

        header_col, status_col = st.columns([4, 1])
        with header_col:
            st.markdown('<div class="section-title">🤖 AI Financial Advisor</div>', unsafe_allow_html=True)
            st.markdown(
                "Your personal AI-powered financial intelligence assistant. Ask anything about your "
                "expenses — it has read all your data and insights."
            )
        with status_col:
            st.markdown(
                '<div style="text-align:right; margin-top:1.6rem;">'
                '<span class="status-pill" style="background:#eef1ff; color:#3a4a99; '
                'border-color:#dbe2ff;"><span class="status-dot status-on"></span>Gemini 2.5 Flash</span>'
                '</div>',
                unsafe_allow_html=True,
            )

        # ── API key status ────────────────────────────────────────────────────
        if not api_key_set:
            st.warning(
                "**Gemini API key not found.** "
                "Create a `.env` file in the project folder with:\n"
                "```\nGEMINI_API_KEY=your_key_here\n```\n"
                "Get a free key at: https://aistudio.google.com/app/apikey"
            )

        # ── Data context panel (collapsible) ──────────────────────────────────
        with st.expander("📋 What the AI knows about your data", expanded=False):
            kc1, kc2, kc3 = st.columns(3)
            kc1.metric("Total Spend",     f"₹{total_spend:,.0f}")
            kc2.metric("Transactions",    f"{total_txns:,}")
            kc3.metric("Health Score",    f"{score}/100 ({label})")
            st.markdown(f"**Top Category:** {top_cat} | **Months:** {n_months} | **Avg Txn:** ₹{avg_txn:,.0f}")
            st.caption(f"{len(insights)} insights loaded into AI context.")

        # ── Suggested questions ───────────────────────────────────────────────
        suggestions = get_suggested_questions(top_cat, label)
        st.markdown("**💬 Suggested questions — click to ask:**")

        # Render suggestion buttons in a 3-column grid
        sug_cols = st.columns(3)
        for idx, suggestion in enumerate(suggestions):
            col = sug_cols[idx % 3]
            if col.button(suggestion, key=f"sug_{idx}", use_container_width=True):
                # Treat click as a user message
                st.session_state._pending_question = suggestion

        # ── Chat message display area ─────────────────────────────────────────
        if st.session_state.chat_messages:
            # Build HTML for all messages
            chat_html = '<div class="chat-container">'
            for msg in st.session_state.chat_messages:
                if msg["role"] == "user":
                    chat_html += (
                        '<div class="chat-row user">'
                        '<div class="chat-name">👤 You</div>'
                        f'<div class="chat-user">{msg["text"]}</div></div>'
                    )
                else:
                    chat_html += (
                        '<div class="chat-row bot">'
                        '<div class="chat-name">🤖 Gemini Advisor</div>'
                        f'<div class="chat-bot">{msg["text"]}</div></div>'
                    )
            chat_html += "</div>"
            st.markdown(chat_html, unsafe_allow_html=True)
        else:
            st.markdown(
                '<div class="chat-container"><div class="chat-empty">'
                '💬 No messages yet.<br>Ask a question below or click a suggestion above.'
                '</div></div>',
                unsafe_allow_html=True,
            )

        # ── Input row: text box + send + clear ───────────────────────────────
        input_col, btn_col, clear_col = st.columns([6, 1, 1])

        with input_col:
            user_input = st.text_input(
                "Ask your financial advisor...",
                key="chat_input",
                placeholder="e.g. How can I reduce my bills spending?",
                label_visibility="collapsed",
            )

        with btn_col:
            send_btn = st.button("Send", use_container_width=True, type="primary")

        with clear_col:
            if st.button("Clear", use_container_width=True):
                st.session_state.chat_messages  = []
                st.session_state.gemini_history = []
                st.rerun()

        # ── Handle suggestion click (from button above) ───────────────────────
        pending = st.session_state.pop("_pending_question", None)
        question_to_send = pending or (user_input.strip() if send_btn and user_input.strip() else None)

        if question_to_send:
            # Add user message to display history
            st.session_state.chat_messages.append({
                "role": "user",
                "text": question_to_send,
            })

            # Get AI response (with spinner)
            with st.spinner("Gemini is thinking..."):
                bot_response = get_chatbot_response(
                    user_input     = question_to_send,
                    insights_list  = insights,
                    summary_stats  = summary_stats,
                    chat_history   = st.session_state.gemini_history,
                )

            # Add bot response to display history
            st.session_state.chat_messages.append({
                "role": "bot",
                "text": bot_response,
            })

            # Update Gemini multi-turn history
            # First turn: includes full context; subsequent turns: just the question
            if not st.session_state.gemini_history:
                from chatbot import build_system_context
                ctx = build_system_context(insights, summary_stats)
                full_first = f"{ctx}\n\n---\nUser Question: {question_to_send}"
                st.session_state.gemini_history.append({"role": "user",  "parts": full_first})
            else:
                st.session_state.gemini_history.append({"role": "user",  "parts": question_to_send})

            st.session_state.gemini_history.append({"role": "model", "parts": bot_response})

            st.rerun()

        # ── Conversation stats footer ─────────────────────────────────────────
        if st.session_state.chat_messages:
            n_turns = len([m for m in st.session_state.chat_messages if m["role"] == "user"])
            st.caption(f"💬 {n_turns} question(s) asked this session  |  Powered by Gemini 2.5 Flash")

# ─────────────────────────────────────────────────────────────────────────────
# Landing page — no file loaded
# ─────────────────────────────────────────────────────────────────────────────
else:
    st.markdown("""
    <div class="hero-card">
        <h1>💰 Smart Expense Analyzer</h1>
        <p>Understand your spending.<br>Improve your financial health.<br>Ask AI for personalized insights.</p>
    </div>
    """, unsafe_allow_html=True)

    feature_cards = [
        ("📊", "Expense Analytics", "Deep-dive metrics on every transaction, category and month."),
        ("❤️", "Financial Health", "A single 0–100 score summarizing your spending habits."),
        ("💡", "Smart Insights", "Automated warnings, savings tips and positive signals."),
        ("📈", "Spending Trends", "Category, monthly and budget-vs-guideline charts."),
        ("🤖", "AI Financial Advisor", "Chat with Gemini about your own expense data."),
        ("📋", "Automated Reports", "Export a full Excel report in one click."),
    ]

    for row_start in range(0, len(feature_cards), 3):
        cols = st.columns(3)
        for col, (icon, title, desc) in zip(cols, feature_cards[row_start:row_start + 3]):
            with col:
                st.markdown(f"""
                <div class="feature-card">
                    <h4>{icon} {title}</h4>
                    <p>{desc}</p>
                </div>
                """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col_a, col_b = st.columns([3, 2])

    with col_a:
        st.info("👈 Your CSV path is pre-filled in the sidebar. Click **Load from Path** to begin.")
        st.markdown("""
        ### Expected CSV Format
        | Date | Exp Type | Amount |
        |---|---|---|
        | 29-Oct-14 | Bills | 82475 |
        | 22-Aug-14 | Food | 32555 |
        """)

    with col_b:
        st.markdown("""
        <div class="feature-card">
            <h4>Quick Start</h4>
            <ol style='color:#333;line-height:2.1; font-size:0.88rem; padding-left:1.1rem;'>
                <li>Install: <code>pip install -r requirements.txt</code></li>
                <li>Add your Gemini API key to a <code>.env</code> file:<br>
                    <code>GEMINI_API_KEY=your_key_here</code></li>
                <li>Run: <code>streamlit run app.py</code></li>
                <li>Click <b>Load from Path</b> in the sidebar</li>
                <li>Go to <b>AI Financial Advisor</b> tab to chat!</li>
            </ol>
            <p style='margin-top:0.6rem;font-size:0.8rem;color:#666;'>
                Get a free Gemini API key at:<br>
                <a href='https://aistudio.google.com/app/apikey' target='_blank'>
                aistudio.google.com/app/apikey</a>
            </p>
        </div>
        """, unsafe_allow_html=True)